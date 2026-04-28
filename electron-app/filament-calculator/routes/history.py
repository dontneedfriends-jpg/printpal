from flask import Blueprint, request, render_template, jsonify, flash, redirect, url_for
from database import get_db
from config import UPLOAD_DIR
import json
import os

history_bp = Blueprint("history", __name__)


@history_bp.route("/history")
def history():
    db = get_db()
    client_filter = request.args.get("client_id", "")
    sql = """
        SELECT c.*, p.name as printer_name, p.power_watts, f.name as filament_name, f.color as filament_color,
               cl.name as client_name
        FROM calculations c
        JOIN printers p ON c.printer_id = p.id
        JOIN filaments f ON c.filament_id = f.id
        LEFT JOIN clients cl ON c.client_id = cl.id
    """
    params = []
    if client_filter:
        sql += " WHERE c.client_id = ?"
        params.append(client_filter)
    sql += " ORDER BY c.created_at DESC"
    calc_list = db.execute(sql, params).fetchall()
    clients = db.execute("SELECT id, name FROM clients ORDER BY name").fetchall()
    db.close()
    return render_template("history.html", calculations=calc_list, clients=clients, client_filter=client_filter, lang=request.lang)


@history_bp.route("/history/<int:id>/delete", methods=["POST"])
def delete_calculation(id):
    db = get_db()
    row = db.execute("SELECT model_file, filament_data FROM calculations WHERE id = ?", (id,)).fetchone()
    if row:
        if row["model_file"]:
            fpath = os.path.join(UPLOAD_DIR, row["model_file"])
            if os.path.exists(fpath):
                os.remove(fpath)
        if row["filament_data"]:
            try:
                filament_data = json.loads(row["filament_data"])
                for f in filament_data:
                    db.execute("UPDATE filaments SET remaining_g = remaining_g + ? WHERE id = ?", (f["weight"], f["id"]))
            except (ValueError, TypeError, KeyError) as e:
                import logging
                logging.getLogger(__name__).warning(f"Failed to restore filament data: {e}")
        db.execute("DELETE FROM calculations WHERE id = ?", (id,))
        db.commit()
    db.close()
    return jsonify({"ok": True})


@history_bp.route("/history/export")
def export_history():
    db = get_db()
    calc_list = db.execute("""
        SELECT c.*, p.name as printer_name, f.name as filament_name, f.color as filament_color
        FROM calculations c
        JOIN printers p ON c.printer_id = p.id
        JOIN filaments f ON c.filament_id = f.id
        ORDER BY c.created_at DESC
    """).fetchall()
    db.close()

    output = __import__('io').StringIO()
    writer = __import__('csv').writer(output, delimiter=';')
    writer.writerow(['Дата', 'Модель', 'Принтер', 'Филамент', 'Цвет', 'Вес (г)', 'Время (ч)', 'Филамент (руб)', 'Электричество (руб)', 'Амортизация (руб)', 'Наценка (%)', 'Итого (руб)'])
    for c in calc_list:
        writer.writerow([
            c["created_at"][:10], c["model_name"], c["printer_name"],
            c["filament_name"], c["filament_color"], "%.1f" % c["weight_g"],
            "%.1f" % c["print_time_hours"], "%.2f" % c["filament_cost"],
            "%.2f" % c["electricity_cost"], "%.2f" % c["depreciation_cost"],
            "%.0f" % c["markup_percent"], "%.2f" % c["total_cost"]
        ])
    output.seek(0)
    from flask import Response
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment;filename=printpal_history.csv"})


@history_bp.route("/history/export/excel")
def export_history_excel():
    db = get_db()
    calc_list = db.execute("""
        SELECT c.*, p.name as printer_name, f.name as filament_name, f.color as filament_color
        FROM calculations c
        JOIN printers p ON c.printer_id = p.id
        JOIN filaments f ON c.filament_id = f.id
        ORDER BY c.created_at DESC
    """).fetchall()
    db.close()

    # Convert sqlite3.Row to plain dict so .get() works
    calc_list = [dict(c) for c in calc_list]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "История расчётов"

        headers = ['Дата', 'Модель', 'Принтер', 'Филамент', 'Цвет', 'Вес (г)', 'Время (ч)', 'Базовая ставка', 'Филамент (руб)', 'Электричество (руб)', 'Амортизация (руб)', 'Иные расходы', 'Наценка (%)', 'Итого (руб)']
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for c in calc_list:
            ws.append([
                c["created_at"][:10] if c["created_at"] else "",
                c["model_name"] or "",
                c["printer_name"] or "",
                c["filament_name"] or "",
                c["filament_color"] or "",
                c["weight_g"] or 0,
                c["print_time_hours"] or 0,
                c["base_rate"] or 0,
                c["filament_cost"] or 0,
                c["electricity_cost"] or 0,
                c["depreciation_cost"] or 0,
                c.get("other_expenses") if c.get("other_expenses") is not None else 0,
                c["markup_percent"] or 0,
                c["total_cost"] or 0,
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import Response
        return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment;filename=printpal_history.xlsx"})
    except ImportError:
        flash("openpyxl не установлен. Используйте CSV экспорт.", "warning")
        return redirect(url_for(".history"))
